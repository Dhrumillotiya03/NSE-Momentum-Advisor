"""
export_sr_log_xlsx.py
----------------------
Formatted .xlsx snapshot of sr_daily_log.csv for sharing with the client
(they track S/R only, no need for the live model/pipeline).

sr_daily_log.csv itself must stay plain CSV — it's overwritten daily by
sr_daily_logger.py, and Excel formatting saved back into a .csv is silently
discarded anyway (CSV is plain text, has no styling). This script instead
writes a SEPARATE .xlsx with bold header, frozen header row, and auto-fit
column widths (esp. Date, which truncates in the raw CSV at default width).

Run whenever you want to refresh the client's copy:
    python export_sr_log_xlsx.py
    python export_sr_log_xlsx.py ../data/sr_dynamic_log.csv   ← export a different log
"""
import sys
import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

SRC_DEFAULT = "../data/sr_daily_log.csv"


def export(src_path):
    df = pd.read_csv(src_path)
    out_path = src_path.rsplit(".", 1)[0] + ".xlsx"

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="SR Log")
        ws = writer.sheets["SR Log"]

        for cell in ws[1]:
            cell.font = Font(bold=True)
        ws.freeze_panes = "A2"

        for i, col in enumerate(df.columns, start=1):
            width = max(len(str(col)), df[col].astype(str).map(len).max()) + 2
            ws.column_dimensions[get_column_letter(i)].width = width

    print(f"✅ Wrote {out_path} ({len(df)} rows)")


if __name__ == "__main__":
    export(sys.argv[1] if len(sys.argv) > 1 else SRC_DEFAULT)
