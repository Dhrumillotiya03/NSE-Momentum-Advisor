"""
run_client_sr_log.py
---------------------
STANDALONE daily S/R runner for the client's copy of the codebase. Client
only tracks support/resistance — does NOT need download_data.py's full
~500-stock universe download, nor exit_engine/paper_trader/agent_sim/etc.
from the main run_daily_log.sh pipeline (those are internal strategy tooling
and shouldn't ship to an external client).

WATCHLIST is read live from sr_daily_logger.py at each run — the client has
already edited their own copy of that list and may keep adding to it, so
this script must never hardcode a symbol count or specific tickers.

HANDOFF: on the client's machine, replace their sr_daily_logger.py with the
contents of client_sr_daily_logger.py (this repo's client-bound reference
copy, kept out of the production pipeline) — file must stay named
sr_daily_logger.py there since that's what this import line below expects.
See client_sr_daily_logger.py's docstring for the full file list that also
needs updating on their machine (support_resistance.py, sr_reach_table.json,
download_index.py, requirements.txt).

Does exactly four things:
  1. Downloads price history for whatever's currently in WATCHLIST (yfinance)
  2. Runs sr_daily_logger.log_stock() for each -> in-memory rows
  3. Writes data/sr_daily_log.xlsx (formatted; NOT .csv — see export_sr_log_xlsx.py
     for why .csv can't hold Excel formatting, and why round-tripping through
     Excel silently discards it) — FULL history, all columns, for month-long
     tracking
  4. Writes data/sr_today.xlsx — TODAY's rows only, trimmed to the columns the
     client actually reads at a glance (Date, Symbol, CMP, S1, S1_prob, S2,
     R1, R1_prob, R2), overwritten fresh each run so there's one small file
     to open instead of scrolling the combined log for today's date

Run this once daily after market close (after 3:30pm IST):
    python run_client_sr_log.py

Requires: yfinance, pandas, openpyxl (pip install -r requirements.txt)
"""
import os
from datetime import datetime
import yfinance as yf
import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from sr_daily_logger import WATCHLIST, log_stock, COLUMNS, merge_log
from support_resistance import INDEX_FILES, INDEX_DIR
import download_index

PRICE_DIR = "../data/price_data/"
ETF_DIR = "../data/etf_data/"
OUT_XLSX = "../data/sr_daily_log.xlsx"
OUT_TODAY_XLSX = "../data/sr_today.xlsx"
TODAY_COLUMNS = ["Date", "Symbol", "CMP", "S1", "S1_prob", "S2", "R1", "R1_prob", "R2"]
START_DATE = "2015-01-01"


def is_etf(sym):
    """Route to etf_data/ (support_resistance.py's fallback dir) instead of
    guessing from a hardcoded symbol list, since the client can add new
    ETF/index-fund tickers to WATCHLIST that we'd never know about ahead of
    time.

    Symbol-keyword check FIRST, no network call: yf.Ticker().info is a
    separate, much slower/flakier request than yf.download() (prone to
    throttling and misleading "possibly delisted" errors on live-traded
    names when it gets rate-limited) — calling it for every plain stock in
    a 20+ symbol watchlist was the actual cause of both the multi-minute
    hangs and the spurious WIPRO error reported 2026-07-20. Only fall
    through to .info if the symbol name gives no hint either way, and even
    then treat a failure as "not an ETF" rather than retrying/hanging."""
    keywords = ("ETF", "BEES", "FUND")
    if any(k in sym.upper() for k in keywords):
        return True
    try:
        info = yf.Ticker(sym).info
        qtype = str(info.get("quoteType", "")).upper()
        name = f"{info.get('longName', '')} {info.get('shortName', '')}".upper()
        return qtype in ("ETF", "MUTUALFUND") or any(k in name for k in keywords)
    except Exception:
        return False


def is_fresh(path):
    if not os.path.exists(path):
        return False
    age_days = (datetime.now().timestamp() - os.path.getmtime(path)) / 86400
    return age_days < 1


def already_fresh(sym):
    """True if a symbol's CSV (either price_data/ or etf_data/) was already
    downloaded today. The client's launcher runs download_index.py and
    download_data.py BEFORE this script — download_data.py already covers
    every plain NSE stock in a 200/500-name universe file, so re-downloading
    them here via yfinance again was pure duplicate work (~1 min wasted on a
    52-symbol watchlist). Only symbols download_data.py doesn't cover
    (ETFs, or tickers outside its universe list) actually need fetching here."""
    return any(is_fresh(f"{d}{sym}.csv") for d in (PRICE_DIR, ETF_DIR))


def download_watchlist():
    os.makedirs(PRICE_DIR, exist_ok=True)
    os.makedirs(ETF_DIR, exist_ok=True)
    stock_syms = [s for s in WATCHLIST if s.upper() not in INDEX_FILES]
    index_syms = [s for s in WATCHLIST if s.upper() in INDEX_FILES]
    if index_syms:
        # NIFTY50/INDIAVIX read from data/index_data/ (support_resistance's
        # INDEX_FILES map), not yfinance directly — but that file still needs
        # refreshing daily, same source download_index.py already uses.
        # Skipped if the client's launcher already ran download_index.py today.
        if is_fresh(f"{INDEX_DIR}nifty50.csv"):
            print("Index data already fresh today — skipping re-download")
        else:
            print(f"Refreshing index data for: {', '.join(index_syms)}")
            download_index.main()

    to_fetch = [s for s in stock_syms if not already_fresh(s)]
    skipped = len(stock_syms) - len(to_fetch)
    print(f"Downloading {len(to_fetch)} watchlist symbols"
          + (f" ({skipped} already fresh today, skipped)" if skipped else "") + "...")
    failed = []
    for sym in to_fetch:
        try:
            # timeout= caps a single stuck request instead of hanging the
            # whole run — without it, one flaky symbol (Yahoo throttling,
            # bad connection) blocks every symbol after it indefinitely.
            df = yf.download(sym, start=START_DATE, interval="1d",
                              auto_adjust=True, progress=False, timeout=20)
        except Exception as e:
            print(f"  ⚠️ {sym}: download error — {e}")
            failed.append(sym)
            continue
        if df.empty:
            print(f"  ⚠️ {sym}: no data returned (check ticker is still valid/listed)")
            failed.append(sym)
            continue
        if isinstance(df.columns, pd.MultiIndex):
            df.columns = df.columns.get_level_values(0)
        df = df.reset_index()
        df["Symbol"] = sym
        out_dir = ETF_DIR if is_etf(sym) else PRICE_DIR
        df.to_csv(f"{out_dir}{sym}.csv", index=False)
    print("Download done.")
    if failed:
        print(f"  {len(failed)} symbol(s) had no data this run (kept using "
              f"yesterday's cached CSV, if any): {', '.join(failed)}\n")
    else:
        print()


def write_xlsx(df, path):
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="SR Log")
        ws = writer.sheets["SR Log"]
        for cell in ws[1]:
            cell.font = Font(bold=True)
        ws.freeze_panes = "A2"
        for i, col in enumerate(df.columns, start=1):
            # NaN/None cells: .astype(str) doesn't reliably coerce these to a
            # plain string across pandas/numpy versions (observed crashing
            # with "float has no len()" on a client machine, Python 3.13 /
            # newer pandas — likely a nullable-dtype code path we don't hit
            # here) — map each value through str() by hand instead, which
            # always returns an actual string ("nan" for a NaN).
            cell_lens = [len(str(v)) for v in df[col]]
            width = max([len(str(col))] + cell_lens) + 2
            ws.column_dimensions[get_column_letter(i)].width = width


def main():
    download_watchlist()

    print(f"Logging S/R snapshot for {len(WATCHLIST)} stocks")
    print("─" * 50)
    rows = []
    for sym in WATCHLIST:
        row = log_stock(sym)
        if row:
            rows.append(row)
            print(f"  ✅ {row['Symbol']:<14} {row['Date']}  CMP ₹{row['CMP']}")

    if not rows:
        print("Nothing logged.")
        return

    new_df = pd.DataFrame(rows, columns=COLUMNS)

    # Keep history across runs the same way sr_daily_logger.py does, but
    # persist it in a hidden CSV cache (not the client-facing xlsx) so
    # accumulated rows survive between days without exposing a .csv to edit.
    cache_path = "../data/.sr_daily_log_cache.csv"
    old_csv_path = "../data/sr_daily_log.csv"
    if not os.path.exists(cache_path) and os.path.exists(old_csv_path):
        # First run of this script on a machine that was previously running
        # sr_daily_logger.py directly: seed the cache from that file's history
        # instead of starting blank, so the client doesn't lose their existing
        # log the day they switch over.
        print(f"[first run] seeding cache from existing {old_csv_path}")
        pd.read_csv(old_csv_path).to_csv(cache_path, index=False)

    combined = merge_log(new_df, cache_path)
    combined.to_csv(cache_path, index=False)

    write_xlsx(combined, OUT_XLSX)
    print("─" * 50)
    print(f"✅ Wrote {OUT_XLSX} ({len(combined)} rows)")

    # Today-only view, trimmed columns: new_df is exactly today's rows (per
    # symbol, its own latest completed-candle date — see log_stock's Date
    # comment), before the merge into combined history above. Overwritten
    # fresh every run so the client always has one small file for "today's
    # levels" instead of hunting for today's date inside the combined log.
    write_xlsx(new_df[TODAY_COLUMNS], OUT_TODAY_XLSX)
    print(f"✅ Wrote {OUT_TODAY_XLSX} ({len(new_df)} rows)")


if __name__ == "__main__":
    main()
